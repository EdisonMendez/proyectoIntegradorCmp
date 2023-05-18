from openpyxl import load_workbook
import pandas as pd
import numpy as np
import matplotlib as plt
from gestionDB.models import ventasHistoricas #tabla a insertar en BD
from collections import defaultdict
import csv

def ventas2017():

    wb2017 = load_workbook("historics/2017.xlsx")
    sh2017 = wb2017["2017"]

    codigo = []
    nombreCodigo = []
    fecha = []
    bodega = []
    clienteNombre = []
    unidadesVendidas = []
    precio = []
    ventaNeta = []

    for index, item in enumerate(sh2017["A2":"A377941"]):
        if index > 0:
            for cell in item:
                codigo.append(str(cell.value))

    for index, item in enumerate(sh2017["B2":"B377941"]):
        if index > 0:
            for cell in item:
                nombreCodigo.append(str(cell.value))

    for index, item in enumerate(sh2017["C2":"C377941"]):
        if index > 0:
            for cell in item:
                fecha.append(str(cell.value))

    for index, item in enumerate(sh2017["F2":"F377941"]):
        if index > 0:
            for cell in item:
                bodega.append(str(cell.value))

    for index, item in enumerate(sh2017["G2":"G377941"]):
        if index > 0:
            for cell in item:
                clienteNombre.append(str(cell.value))

    for index, item in enumerate(sh2017["M2":"M377941"]):
        if index > 0:
            for cell in item:
                unidadesVendidas.append(str(int(abs(cell.value))))

    for index, item in enumerate(sh2017["N2":"N377941"]):
        if index > 0:
            for cell in item:
                precio.append(str(cell.value))

    for index, item in enumerate(sh2017["O2":"O377941"]):
        if index > 0:
            for cell in item:
                ventaNeta.append(str(cell.value))



    diccionario = {'Codigo': codigo,
                    'Nombre Codigo': nombreCodigo,
                   'Fecha Venta': fecha, 'bodega' : bodega, 'Nombre cliente' : clienteNombre,'Unidades_vendidas' : unidadesVendidas,
                   'precio' : precio, 'Venta (dolares)' : ventaNeta
                   }



    dataFrame = pd.DataFrame(diccionario)

    dfCopy = dataFrame.groupby(['Codigo','Fecha Venta'])['Unidades_vendidas'].agg(['sum','count'])
    dfCopy.to_csv('converted/ventasPorFecha.csv', index=False)
    print("Done.")


    dataFrame = dataFrame[dataFrame.Unidades_vendidas != 0]
    dataFrame.to_csv('converted/ventas2017.csv', index=False)
    dataFrame = dataFrame[(dataFrame != 0).all(axis=1)]


    ####insert to DB
    diccionario1 = dict(zip(codigo, codigo))
    diccionario2 = dict(zip(codigo, nombreCodigo))
    diccionario3 = dict(zip(codigo, fecha))
    diccionario4 = dict(zip(codigo, bodega))
    diccionario5 = dict(zip(codigo, clienteNombre))
    diccionario6 = dict(zip(codigo, unidadesVendidas))
    diccionario7 = dict(zip(codigo, precio))
    diccionario8 = dict(zip(codigo, ventaNeta))

    for i in diccionario1:
         sql = ventasHistoricas.objects.create(codigo=diccionario1.get(i,i),nombreCodigo=diccionario2.get(i,i),
                                               fecha=diccionario3.get(i,i)
                                               ,bodega=diccionario4.get(i,i),clienteNombre=diccionario5.get(i,i),
                                               unidadesVendidas=diccionario6.get(i,i),precio=diccionario7.get(i,i)
                                               ,ventaNeta=diccionario8.get(i,i)
                                               )

    print("Info cargada exitosamente a la base de datos.")



def ventas2018():

    wb2018 = load_workbook("historics/2018.xlsx")
    sh2018 = wb2018["2018"]

    codigo = []
    nombreCodigo = []
    fecha = []
    bodega = []
    clienteNombre = []
    unidadesVendidas = []
    precio = []
    ventaNeta = []

    for index, item in enumerate(sh2018["A2":"A413031"]):
        if index > 0:
            for cell in item:
                codigo.append(str(cell.value))

    for index, item in enumerate(sh2018["B2":"B413031"]):
        if index > 0:
            for cell in item:
                nombreCodigo.append(str(cell.value))

    for index, item in enumerate(sh2018["C2":"C413031"]):
        if index > 0:
            for cell in item:
                fecha.append(str(cell.value))

    for index, item in enumerate(sh2018["F2":"F413031"]):
        if index > 0:
            for cell in item:
                bodega.append(str(cell.value))

    for index, item in enumerate(sh2018["G2":"G413031"]):
        if index > 0:
            for cell in item:
                clienteNombre.append(str(cell.value))

    for index, item in enumerate(sh2018["M2":"M413031"]):
        if index > 0:
            for cell in item:
                unidadesVendidas.append(str(abs(cell.value)))

    for index, item in enumerate(sh2018["N2":"N413031"]):
        if index > 0:
            for cell in item:
                precio.append(str(cell.value))

    for index, item in enumerate(sh2018["O2":"O413031"]):
        if index > 0:
            for cell in item:
                ventaNeta.append(str(cell.value))



    diccionario = {'Codigo': codigo,
                    'Nombre Codigo': nombreCodigo,
                   'Fecha Venta': fecha, 'bodega' : bodega, 'Nombre cliente' : clienteNombre,'Unidades_vendidas' : unidadesVendidas,
                   'precio' : precio, 'Venta (dolares)' : ventaNeta
                   }

    dataFrame = pd.DataFrame(diccionario)
    dataFrame = dataFrame[(dataFrame != 0).all(axis=1)]


    dataFrame.to_csv('converted/ventas2018.csv',index=False)

def ventas2019():
    wb2019 = load_workbook("historics/2019.xlsx")
    sh2019 = wb2019["2019"]

    codigo = []
    nombreCodigo = []
    fecha = []
    bodega = []
    clienteNombre = []
    unidadesVendidas = []
    precio = []
    ventaNeta = []

    for index, item in enumerate(sh2019["A2":"A445166"]):
        if index > 0:
            for cell in item:
                codigo.append(str(cell.value))

    for index, item in enumerate(sh2019["B2":"B445166"]):
        if index > 0:
            for cell in item:
                nombreCodigo.append(str(cell.value))

    for index, item in enumerate(sh2019["C2":"C445166"]):
        if index > 0:
            for cell in item:
                fecha.append(str(cell.value))

    for index, item in enumerate(sh2019["F2":"F445166"]):
        if index > 0:
            for cell in item:
                bodega.append(str(cell.value))

    for index, item in enumerate(sh2019["G2":"G445166"]):
        if index > 0:
            for cell in item:
                clienteNombre.append(str(cell.value))

    for index, item in enumerate(sh2019["M2":"M445166"]):
        if index > 0:
            for cell in item:
                unidadesVendidas.append(str(abs(cell.value)))

    for index, item in enumerate(sh2019["N2":"N445166"]):
        if index > 0:
            for cell in item:
                precio.append(str(cell.value))

    for index, item in enumerate(sh2019["O2":"O445166"]):
        if index > 0:
            for cell in item:
                ventaNeta.append(str(cell.value))

    diccionario = {'Codigo': codigo,
                   'Nombre Codigo': nombreCodigo,
                   'Fecha Venta': fecha, 'bodega': bodega, 'Nombre cliente': clienteNombre,
                   'Unidades_vendidas': unidadesVendidas,
                   'precio': precio, 'Venta (dolares)': ventaNeta
                   }

    dataFrame = pd.DataFrame(diccionario)

    dataFrame.to_csv('converted/ventas2019.csv', index=False)

def ventas2020():
    wb2020 = load_workbook("historics/2020.xlsx")
    sh2020 = wb2020["2020"]

    codigo = []
    nombreCodigo = []
    fecha = []
    bodega = []
    clienteNombre = []
    unidadesVendidas = []
    precio = []
    ventaNeta = []

    for index, item in enumerate(sh2020["A2":"A428340"]):
        if index > 0:
            for cell in item:
                codigo.append(str(cell.value))

    for index, item in enumerate(sh2020["B2":"B428340"]):
        if index > 0:
            for cell in item:
                nombreCodigo.append(str(cell.value))

    for index, item in enumerate(sh2020["C2":"C428340"]):
        if index > 0:
            for cell in item:
                fecha.append(str(cell.value))

    for index, item in enumerate(sh2020["F2":"F428340"]):
        if index > 0:
            for cell in item:
                bodega.append(str(cell.value))

    for index, item in enumerate(sh2020["G2":"G428340"]):
        if index > 0:
            for cell in item:
                clienteNombre.append(str(cell.value))

    for index, item in enumerate(sh2020["M2":"M428340"]):
        if index > 0:
            for cell in item:
                unidadesVendidas.append(str(abs(cell.value)))

    for index, item in enumerate(sh2020["N2":"N428340"]):
        if index > 0:
            for cell in item:
                precio.append(str(cell.value))

    for index, item in enumerate(sh2020["O2":"O428340"]):
        if index > 0:
            for cell in item:
                ventaNeta.append(str(cell.value))

    diccionario = {'Codigo': codigo,
                   'Nombre Codigo': nombreCodigo,
                   'Fecha Venta': fecha, 'bodega': bodega, 'Nombre cliente': clienteNombre,
                   'Unidades_vendidas': unidadesVendidas,
                   'precio': precio, 'Venta (dolares)': ventaNeta
                   }

    dataFrame = pd.DataFrame(diccionario)


    dataFrame.to_csv('converted/ventas2020.csv', index=False)

def ventas2021():
    wb2021 = load_workbook("historics/2021.xlsx")
    sh2021 = wb2021["2021"]

    codigo = []
    nombreCodigo = []
    fecha = []
    bodega = []
    clienteNombre = []
    unidadesVendidas = []
    precio = []
    ventaNeta = []

    for index, item in enumerate(sh2021["A2":"A506903"]):
        if index > 0:
            for cell in item:
                codigo.append(str(cell.value))

    for index, item in enumerate(sh2021["B2":"B506903"]):
        if index > 0:
            for cell in item:
                nombreCodigo.append(str(cell.value))

    for index, item in enumerate(sh2021["C2":"C506903"]):
        if index > 0:
            for cell in item:
                fecha.append(str(cell.value))

    for index, item in enumerate(sh2021["F2":"F506903"]):
        if index > 0:
            for cell in item:
                bodega.append(str(cell.value))

    for index, item in enumerate(sh2021["G2":"G506903"]):
        if index > 0:
            for cell in item:
                clienteNombre.append(str(cell.value))

    for index, item in enumerate(sh2021["M2":"M506903"]):
        if index > 0:
            for cell in item:
                unidadesVendidas.append(str(int(abs(cell.value))))

    for index, item in enumerate(sh2021["N2":"N506903"]):
        if index > 0:
            for cell in item:
                precio.append(str(cell.value))

    for index, item in enumerate(sh2021["O2":"O506903"]):
        if index > 0:
            for cell in item:
                ventaNeta.append(str(cell.value))

    diccionario = {'Codigo': codigo,
                   'Nombre Codigo': nombreCodigo,
                   'Fecha Venta': fecha, 'bodega': bodega, 'Nombre cliente': clienteNombre,
                   'Unidades_vendidas': unidadesVendidas,
                   'precio': precio, 'Venta (dolares)': ventaNeta
                   }

    dataFrame = pd.DataFrame(diccionario)


    dataFrame.to_csv('converted/ventas2021.csv', index=False)

def ventas2022():

    wb2022 = load_workbook("historics/2022.xlsx")
    sh2022 = wb2022["2022"]

    codigo = []
    nombreCodigo = []
    fecha = []
    bodega = []
    clienteNombre = []
    unidadesVendidas = []
    precio = []
    ventaNeta = []

    for index, item in enumerate(sh2022["A2":"A506903"]):
        if index > 0:
            for cell in item:
                codigo.append(str(cell.value))

    for index, item in enumerate(sh2022["B2":"B506903"]):
        if index > 0:
            for cell in item:
                nombreCodigo.append(str(cell.value))

    for index, item in enumerate(sh2022["C2":"C506903"]):
        if index > 0:
            for cell in item:
                fecha.append(str(cell.value))

    for index, item in enumerate(sh2022["F2":"F506903"]):
        if index > 0:
            for cell in item:
                bodega.append(str(cell.value))

    for index, item in enumerate(sh2022["G2":"G506903"]):
        if index > 0:
            for cell in item:
                clienteNombre.append(str(cell.value))

    for index, item in enumerate(sh2022["M2":"M506903"]):
        if index > 0:
            for cell in item:
                unidadesVendidas.append(str(int(abs(cell.value))))

    for index, item in enumerate(sh2022["N2":"N506903"]):
        if index > 0:
            for cell in item:
                precio.append(str(cell.value))

    for index, item in enumerate(sh2022["O2":"O506903"]):
        if index > 0:
            for cell in item:
                ventaNeta.append(str(cell.value))

    diccionario = {'Codigo': codigo,
                   'Nombre Codigo': nombreCodigo,
                   'Fecha Venta': fecha, 'bodega': bodega, 'Nombre cliente': clienteNombre,
                   'Unidades_vendidas': unidadesVendidas,
                   'precio': precio, 'Venta (dolares)': ventaNeta
                   }

    dataFrame = pd.DataFrame(diccionario)

    # dataFrame.drop(dataFrame.columns[0])
    dataFrame.to_csv('converted/ventas2022.csv', index=False)


def mergeCSVs():
    df = pd.DataFrame()

    csv_files = ['converted/ventas2017.csv', 'converted/ventas2018.csv', 'converted/ventas2019.csv',
                 'converted/ventas2020.csv',
                 'converted/ventas2021.csv', 'converted/ventas2022.csv'
                 ]

    for file in csv_files:
        temp_df = pd.read_csv(file)
        df = pd.concat([df, temp_df])

    df.to_csv('converted/historicsSalesSV.csv', index=False)
    print("Datos unidos correctamente.")









ventas2017()
print('Done 2017')
print()
print('Done 2018')
ventas2018()
print()
print('Done 2019')
ventas2019()
print()
print('Done 2020')
ventas2020()
print()
print('Done 2021')
ventas2021()
print()
print('2022')
ventas2022()

mergeCSVs()